import openpyxl
import json
import os
import re

# File paths
warehouse_json_path = "/Users/bigt_mbair/Library/CloudStorage/GoogleDrive-titankou2002@gmail.com/我的雲端硬碟/BT/Antigravity/倉庫全集/parsed_inventory.json"
enriched_json_path = "/Users/bigt_mbair/Library/CloudStorage/GoogleDrive-titankou2002@gmail.com/我的雲端硬碟/BT/Antigravity/倉庫全集/parsed_inventory_enriched.json"

company_files = {
    "高雅瓷": "/Users/bigt_mbair/Library/CloudStorage/GoogleDrive-titankou2002@gmail.com/我的雲端硬碟/BT/Antigravity/客戶版機器人/高雅瓷內部管理.xlsx",
    "安帝嘉": "/Users/bigt_mbair/Library/CloudStorage/GoogleDrive-titankou2002@gmail.com/我的雲端硬碟/BT/B-T資料/雲端共用試算表/公司資料/安帝嘉-內部管理.xlsx",
    "喜悅納": "/Users/bigt_mbair/Library/CloudStorage/GoogleDrive-titankou2002@gmail.com/我的雲端硬碟/BT/B-T資料/雲端共用試算表/公司資料/喜悅納-內部管理.xlsx"
}

def clean_str(val):
    if val is None:
        return ""
    return str(val).strip()

def clean_hanhwa_code(val):
    if val is None:
        return ""
    val_str = str(val).strip()
    if val_str.endswith(".0"):
        val_str = val_str[:-2]
    return val_str

def safe_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        cleaned = re.sub(r'[^\d.-]', '', str(v))
        return float(cleaned) if cleaned else 0.0
    except ValueError:
        return 0.0

def estimate_pallet_specs(size_str, kg_per_box):
    if not size_str:
        return 24, (24 * kg_per_box if kg_per_box else 600.0)
    
    clean_size = str(size_str).lower().replace("cm", "").replace(" ", "")
    nums = re.findall(r'\d+(?:\.\d+)?', clean_size)
    if len(nums) >= 2:
        try:
            w = float(nums[0])
            h = float(nums[1])
            d1, d2 = sorted([w, h])
        except ValueError:
            return 24, (24 * kg_per_box if kg_per_box else 600.0)
    else:
        return 24, (24 * kg_per_box if kg_per_box else 600.0)
        
    # Estimate boxes based on tile size
    if d2 >= 100:
        boxes = 20
    elif d1 >= 60 and d2 >= 120:
        boxes = 24
    elif d1 >= 60 and d2 >= 60:
        boxes = 32
    elif d2 >= 75:
        boxes = 44
    elif d2 >= 60:
        boxes = 48
    else:
        boxes = 72
        
    kg_per_box_val = float(kg_per_box) if kg_per_box is not None else 25.0
    kg_per_pallet = boxes * kg_per_box_val
    return boxes, round(kg_per_pallet, 1)

def match_batches(b1, b2):
    clean1 = str(b1).strip().upper().replace("。", "")
    clean2 = str(b2).strip().upper().replace("。", "")
    if not clean1 or not clean2:
        return False
    
    if clean1 == clean2:
        return True
        
    # Get alphanumeric prefix (e.g., GH4 from GH4P/8)
    prefix1 = re.match(r'^[A-Z0-9]+', clean1)
    prefix2 = re.match(r'^[A-Z0-9]+', clean2)
    if prefix1 and prefix2:
        p1 = prefix1.group(0)
        p2 = prefix2.group(0)
        if len(p1) >= 2 and len(p2) >= 2:
            if p1 == p2 or p1.startswith(p2) or p2.startswith(p1):
                return True
                
    return clean1 in clean2 or clean2 in clean1

def main():
    print("Starting data enrichment pipeline...")
    
    # 1. Load warehouse inventory
    if not os.path.exists(warehouse_json_path):
        print(f"Error: Warehouse JSON file not found at {warehouse_json_path}")
        return
        
    with open(warehouse_json_path, "r", encoding="utf-8") as f:
        warehouse_pallets = json.load(f)
    print(f"Loaded {len(warehouse_pallets)} pallets from warehouse inventory.")

    # 2. Parse price list and stock lists from the three branches
    price_database = {} # SKU -> list of price entries
    stock_database = {} # SKU -> list of stock entries
    
    for company, path in company_files.items():
        print(f"\nProcessing Excel for company: {company} ({os.path.basename(path)})...")
        if not os.path.exists(path):
            print(f"Warning: File not found at {path}, skipping.")
            continue
            
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        
        # A. Parse price list
        if "編號價目表" in wb.sheetnames:
            price_sheet = wb["編號價目表"]
            rows = list(price_sheet.iter_rows(values_only=True))
            if rows:
                headers = [clean_str(h) for h in rows[0]]
                # Map col indexes
                brand_idx = headers.index("廠牌") if "廠牌" in headers else 0
                origin_idx = headers.index("產地") if "產地" in headers else 1
                series_idx = headers.index("系列") if "系列" in headers else 2
                name_idx = headers.index("原廠品名") if "原廠品名" in headers else 3
                
                # Check for "中文系列" or second "系列"
                cn_series_idx = -1
                if "中文系列" in headers:
                    cn_series_idx = headers.index("中文系列")
                else:
                    indices = [i for i, h in enumerate(headers) if h == "系列"]
                    if len(indices) > 1:
                        cn_series_idx = indices[1]
                    elif len(indices) == 1:
                        cn_series_idx = indices[0]
                        
                sku_idx = headers.index("編號") if "編號" in headers else 5
                size_idx = headers.index("尺寸(cm)") if "尺寸(cm)" in headers else 7
                pcs_idx = headers.index("片/箱") if "片/箱" in headers else 9
                kg_idx = headers.index("KG/箱") if "KG/箱" in headers else 10
                price_idx = headers.index("同行價") if "同行價" in headers else 13
                hanhwa_idx = headers.index("漢樺編號") if "漢樺編號" in headers else 15
                image_idx = headers.index("雲端圖片") if "雲端圖片" in headers else -1
                
                # Check for "單片連結網址" or any column containing "單片"
                single_piece_idx = -1
                if "單片連結網址" in headers:
                    single_piece_idx = headers.index("單片連結網址")
                else:
                    for idx, h in enumerate(headers):
                        if "單片" in h:
                            single_piece_idx = idx
                            break
                
                for r in rows[1:]:
                    sku = clean_str(r[sku_idx])
                    if not sku:
                        continue
                    
                    kg_val = safe_float(r[kg_idx])
                    pcs_val = safe_float(r[pcs_idx])
                    size_val = clean_str(r[size_idx])
                    hanhwa_code = clean_hanhwa_code(r[hanhwa_idx]) if hanhwa_idx < len(r) else ""
                    
                    # Estimate pallet specs
                    boxes_per_pallet, kg_per_pallet = estimate_pallet_specs(size_val, kg_val)
                    
                    price_entry = {
                        "StandardSKU": sku,
                        "HanhwaCode": hanhwa_code,
                        "Company": company,
                        "Brand": clean_str(r[brand_idx]) if brand_idx < len(r) else "",
                        "Origin": clean_str(r[origin_idx]) if origin_idx < len(r) else "",
                        "Series": clean_str(r[series_idx]) if series_idx < len(r) else "",
                        "OriginalName": clean_str(r[name_idx]) if name_idx < len(r) else "",
                        "ChineseSeries": clean_str(r[cn_series_idx]) if cn_series_idx > -1 and cn_series_idx < len(r) else "",
                        "Size": size_val,
                        "PiecesPerBox": pcs_val,
                        "KgPerBox": kg_val,
                        "BoxesPerPallet": boxes_per_pallet,
                        "KgPerPallet": kg_per_pallet,
                        "Price": clean_str(r[price_idx]) if price_idx < len(r) else "",
                        "CloudImage": clean_str(r[image_idx]) if image_idx > -1 and image_idx < len(r) else "",
                        "SinglePieceImage": clean_str(r[single_piece_idx]) if single_piece_idx > -1 and single_piece_idx < len(r) else ""
                    }
                    
                    # Store standard SKU mapping (allow multiple entries per SKU)
                    if sku not in price_database:
                        price_database[sku] = []
                    
                    # Prevent duplicate mappings within the same company
                    if not any(pe["Company"] == company for pe in price_database[sku]):
                        price_database[sku].append(price_entry)
                    
                    # Also map using Hanhwa code if present
                    if hanhwa_code:
                        if hanhwa_code not in price_database:
                            price_database[hanhwa_code] = []
                        if not any(pe["Company"] == company for pe in price_database[hanhwa_code]):
                            price_database[hanhwa_code].append(price_entry)
                        
            print(f"  Loaded {len(price_database)} price mapping groups.")
            
        # B. Parse stock list
        if "庫存表" in wb.sheetnames:
            stock_sheet = wb["庫存表"]
            rows = list(stock_sheet.iter_rows(values_only=True))
            if rows:
                # We know columns by exact index match
                # 0: SKU, 1: Name, 3: Remarks, 8: Pieces, 9: ActualSize, 10: Batch, 11: NominalSize, 12: Pyeongs
                for r_idx, r in enumerate(rows[1:]):
                    sku = clean_str(r[0])
                    if not sku:
                        continue
                    
                    name_val = clean_str(r[1])
                    if "小計" in name_val or "合計" in name_val:
                        continue
                        
                    pieces_val = safe_float(r[8]) if len(r) > 8 else 0.0
                    actual_size = clean_str(r[9]) if len(r) > 9 else ""
                    batch_val = clean_str(r[10]) if len(r) > 10 else ""
                    nominal_size = clean_str(r[11]) if len(r) > 11 else ""
                    remarks = clean_str(r[3]) if len(r) > 3 else ""
                    
                    stock_entry = {
                        "Company": company,
                        "Pieces": pieces_val,
                        "Batch": batch_val if batch_val else "無批號",
                        "ActualSize": actual_size,
                        "NominalSize": nominal_size,
                        "Remarks": remarks
                    }
                    
                    if sku not in stock_database:
                        stock_database[sku] = []
                    stock_database[sku].append(stock_entry)
                    
            print(f"  Loaded {len(stock_database)} stock records.")

    # 3. Enrich the warehouse pallets
    enriched_pallets = []
    skipped_metadata_count = 0
    
    for p in warehouse_pallets:
        sku = clean_str(p.get("SKU"))
        batch = clean_str(p.get("Batch"))
        
        # Initialize default shared metadata
        meta = {
            "Brand": "",
            "Origin": "",
            "Series": "",
            "OriginalName": "",
            "ChineseSeries": "",
            "Size": "",
            "PiecesPerBox": None,
            "KgPerBox": None,
            "BoxesPerPallet": None,
            "KgPerPallet": None,
            "HanhwaCode": "",
            "StandardSKU": "",
            "CloudImage": "",
            "SinglePieceImage": "",
            "Branches": [] # List of company-specific details
        }
        
        # Match by standard SKU or Hanhwa code
        clean_sku_code = clean_hanhwa_code(sku)
        p_records = None
        
        if sku in price_database:
            p_records = price_database[sku]
        elif clean_sku_code in price_database:
            p_records = price_database[clean_sku_code]
            
        if p_records:
            # Use the first record to fill out shared product metadata
            first = p_records[0]
            meta.update({
                "Brand": first["Brand"],
                "Origin": first["Origin"],
                "Series": first["Series"],
                "OriginalName": first["OriginalName"],
                "ChineseSeries": first["ChineseSeries"],
                "Size": first["Size"],
                "PiecesPerBox": first["PiecesPerBox"],
                "KgPerBox": first["KgPerBox"],
                "BoxesPerPallet": first["BoxesPerPallet"],
                "KgPerPallet": first["KgPerPallet"],
                "HanhwaCode": first["HanhwaCode"],
                "StandardSKU": first["StandardSKU"],
                "CloudImage": first["CloudImage"],
                "SinglePieceImage": first["SinglePieceImage"]
            })
            
            # Map standard SKU
            standard_sku = first["StandardSKU"]
            
            # Populate details for each branch
            for pr in p_records:
                comp = pr["Company"]
                price = pr["Price"]
                
                branch_detail = {
                    "Company": comp,
                    "Price": price,
                    "CloudImage": pr["CloudImage"],
                    "SinglePieceImage": pr["SinglePieceImage"],
                    "StockTotalPieces": 0.0,
                    "StockOfficialBatch": "無庫存批號",
                    "BatchMatchesStock": False
                }
                
                # Fetch stock levels for this standard SKU in this company
                if standard_sku in stock_database:
                    entries = [e for e in stock_database[standard_sku] if e["Company"] == comp]
                    total_pieces = sum(e["Pieces"] for e in entries)
                    branch_detail["StockTotalPieces"] = total_pieces
                    
                    # Match batch
                    matched_entry = None
                    for e in entries:
                        if match_batches(batch, e["Batch"]):
                            matched_entry = e
                            break
                            
                    if matched_entry:
                        branch_detail["StockOfficialBatch"] = matched_entry["Batch"]
                        branch_detail["BatchMatchesStock"] = True
                    elif entries:
                        batches = [e["Batch"] for e in entries if e["Batch"]]
                        branch_detail["StockOfficialBatch"] = ", ".join(set(batches)) if batches else "無庫存批號"
                        branch_detail["BatchMatchesStock"] = False
                
                meta["Branches"].append(branch_detail)
        else:
            skipped_metadata_count += 1
            standard_sku = sku
            
            # As fallback, check if it exists in stock database directly
            if standard_sku in stock_database:
                entries = stock_database[standard_sku]
                # Find unique companies in stock entries
                comps = set(e["Company"] for e in entries)
                
                # Fallback estimate pallet specs based on nominal size in stock
                nom_size = ""
                for e in entries:
                    if e["NominalSize"]:
                        nom_size = e["NominalSize"]
                        break
                boxes_per_pallet, kg_per_pallet = estimate_pallet_specs(nom_size, 25.0)
                meta["Size"] = nom_size
                meta["BoxesPerPallet"] = boxes_per_pallet
                meta["KgPerPallet"] = kg_per_pallet
                
                for comp in comps:
                    comp_entries = [e for e in entries if e["Company"] == comp]
                    total_pieces = sum(e["Pieces"] for e in comp_entries)
                    
                    branch_detail = {
                        "Company": comp,
                        "Price": "無牌價",
                        "CloudImage": "",
                        "SinglePieceImage": "",
                        "StockTotalPieces": total_pieces,
                        "StockOfficialBatch": "無庫存批號",
                        "BatchMatchesStock": False
                    }
                    
                    # Match batch
                    matched_entry = None
                    for e in comp_entries:
                        if match_batches(batch, e["Batch"]):
                            matched_entry = e
                            break
                            
                    if matched_entry:
                        branch_detail["StockOfficialBatch"] = matched_entry["Batch"]
                        branch_detail["BatchMatchesStock"] = True
                    elif comp_entries:
                        batches = [e["Batch"] for e in comp_entries if e["Batch"]]
                        branch_detail["StockOfficialBatch"] = ", ".join(set(batches)) if batches else "無庫存批號"
                        branch_detail["BatchMatchesStock"] = False
                        
                    meta["Branches"].append(branch_detail)
        
        # Merge warehouse pallet data with metadata
        enriched_pallet = p.copy()
        enriched_pallet.update(meta)
        
        # If there are no matching branches found, add a fallback branch
        if not enriched_pallet["Branches"]:
            enriched_pallet["Branches"].append({
                "Company": "未知分公司",
                "Price": "",
                "CloudImage": "",
                "SinglePieceImage": "",
                "StockTotalPieces": 0.0,
                "StockOfficialBatch": "無庫存批號",
                "BatchMatchesStock": False
            })
            
        enriched_pallets.append(enriched_pallet)
        
    print(f"\nEnrichment complete.")
    print(f"Skipped price metadata lookup for {skipped_metadata_count} pallets (SKU not in price lists).")
    
    # Calculate how many have at least one matched branch batch
    matched_batches_count = sum(1 for ep in enriched_pallets if any(b["BatchMatchesStock"] for b in ep["Branches"]))
    print(f"Successfully matched batch numbers for {matched_batches_count} pallets.")
    
    # Save enriched data
    with open(enriched_json_path, "w", encoding="utf-8") as f:
        json.dump(enriched_pallets, f, indent=2, ensure_ascii=False)
    print(f"Saved enriched JSON database to: {enriched_json_path}")

if __name__ == "__main__":
    main()
