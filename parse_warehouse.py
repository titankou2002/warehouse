import openpyxl
import json
import re

def identify_bg_color(fill_obj):
    if not fill_obj or not hasattr(fill_obj, 'start_color'):
        return "WHITE"
    color_obj = fill_obj.start_color
    if not color_obj or color_obj.type != 'rgb' or not color_obj.rgb:
        return "WHITE"
    
    rgb_str = color_obj.rgb
    hex_color = rgb_str[-6:].upper()
    if hex_color in ["000000", "FFFFFF"]:
        return "WHITE"
    
    try:
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
    except ValueError:
        return "WHITE"
    
    # Check for Yellow (Red and Green are high, Blue is low)
    if r > 180 and g > 180 and b < 160:
        return "YELLOW"
    # Check for Red (Red is high, Green and Blue are low)
    if r > 150 and r > g * 1.3 and r > b * 1.3:
        return "RED"
    # Check for Green (Green is high, Red and Blue are low)
    if g > 120 and g > r * 1.2 and g > b * 1.2:
        return "GREEN"
    # Check for Purple
    if r > 120 and b > 120 and g < 100:
        return "PURPLE"
        
    return "WHITE"

def identify_font_color(font_obj):
    if not font_obj or not hasattr(font_obj, 'color') or not font_obj.color:
        return "BLACK"
    color_obj = font_obj.color
    if not color_obj or color_obj.type != 'rgb' or not color_obj.rgb:
        return "BLACK"
    
    rgb_str = color_obj.rgb
    hex_color = rgb_str[-6:].upper()
    if hex_color == "000000":
        return "BLACK"
    
    try:
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
    except ValueError:
        return "BLACK"
    
    if b > 150 and b > r * 1.3 and b > g * 1.3:
        return "BLUE"
    if r > 150 and r > g * 1.3 and r > b * 1.3:
        return "RED"
    if g > 120 and g > r * 1.2 and g > b * 1.2:
        return "GREEN"
    
    return "BLACK"

def clean_value(val):
    if val is None:
        return None
    val_str = str(val).strip()
    if val_str in ["None", ""]:
        return None
    return val_str

def parse_sheet(sheet, sheet_name):
    print(f"Parsing sheet: {sheet_name}...")
    
    # Make a copy of merged ranges because we will unmerge them
    merged_ranges = list(sheet.merged_cells.ranges)
    for r in merged_ranges:
        try:
            top_left_cell = sheet.cell(row=r.min_row, column=r.min_col)
            val = top_left_cell.value
            fill = top_left_cell.fill
            font = top_left_cell.font
            alignment = top_left_cell.alignment
            
            sheet.unmerge_cells(start_row=r.min_row, start_column=r.min_col, end_row=r.max_row, end_column=r.max_col)
            
            for row in range(r.min_row, r.max_row + 1):
                for col in range(r.min_col, r.max_col + 1):
                    c = sheet.cell(row=row, column=col)
                    c.value = val
                    c.fill = fill
                    c.font = font
                    c.alignment = alignment
        except Exception as e:
            pass

    pallets = []
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    # Row 1 has slot indices. Find all column pairs:
    col_pairs = []
    for col in range(2, max_col + 1):
        cell_val = clean_value(sheet.cell(row=1, column=col).value)
        if cell_val is not None and cell_val != "數量":
            qty_col = col + 1
            col_pairs.append((col, qty_col, cell_val))
            
    print(f"  Found {len(col_pairs)} slots (columns) to parse.")
    
    # Loop over each slot column
    for prod_col, qty_col, slot_id in col_pairs:
        current_stack = None
        r = 2
        while r <= max_row:
            cell_a = sheet.cell(row=r, column=1)
            cell_prod = sheet.cell(row=r, column=prod_col)
            cell_qty = sheet.cell(row=r, column=qty_col)
            
            val_prod = clean_value(cell_prod.value)
            
            # Check if this cell is a stack header
            if val_prod and "排" in val_prod:
                current_stack = val_prod
                r += 1
                continue
                
            if current_stack is None:
                r += 1
                continue
                
            bg_color = identify_bg_color(cell_prod.fill)
            
            if bg_color != "WHITE" and val_prod is not None:
                # We found a pallet! It might span 1, 2 or more rows of the same background color.
                pallet_rows = []
                temp_r = r
                while temp_r <= max_row:
                    temp_cell_prod = sheet.cell(row=temp_r, column=prod_col)
                    temp_bg = identify_bg_color(temp_cell_prod.fill)
                    if temp_bg == bg_color and clean_value(temp_cell_prod.value) is not None:
                        pallet_rows.append(temp_r)
                        temp_r += 1
                    else:
                        break
                
                sku = None
                batch = None
                box_qty = 0.0
                piece_qty = 0.0
                remarks = []
                
                sku = clean_value(sheet.cell(row=pallet_rows[0], column=prod_col).value)
                
                if len(pallet_rows) > 1:
                    batch = clean_value(sheet.cell(row=pallet_rows[1], column=prod_col).value)
                
                if len(pallet_rows) > 2:
                    for extra_r in pallet_rows[2:]:
                        extra_val = clean_value(sheet.cell(row=extra_r, column=prod_col).value)
                        if extra_val:
                            remarks.append(extra_val)
                            
                for pr in pallet_rows:
                    row_header = clean_value(sheet.cell(row=pr, column=1).value)
                    qty_val = clean_value(sheet.cell(row=pr, column=qty_col).value)
                    
                    if qty_val is not None:
                        try:
                            qty_num = float(qty_val)
                        except ValueError:
                            qty_num = 0.0
                            remarks.append(f"Qty parse error: {qty_val}")
                            
                        if row_header == "編號/箱":
                            box_qty += qty_num
                        elif row_header == "批號/片":
                            piece_qty += qty_num
                        else:
                            piece_qty += qty_num
                
                font_color = identify_font_color(cell_prod.font)
                
                status = "正常庫存"
                if bg_color in ["RED", "YELLOW"]:
                    status = "混板/散板"
                elif bg_color == "GREEN":
                    status = "專案庫存"
                    
                is_last_pallet = False
                if status == "混板/散板" and font_color == "BLUE":
                    is_last_pallet = True
                    
                pallets.append({
                    "Sheet": sheet_name,
                    "Slot": slot_id,
                    "Stack": current_stack,
                    "SKU": sku,
                    "Batch": batch if batch else "無批號",
                    "BoxQty": box_qty,
                    "PieceQty": piece_qty,
                    "BgColor": bg_color,
                    "FontColor": font_color,
                    "Status": status,
                    "IsLastPallet": is_last_pallet,
                    "Remarks": ", ".join(remarks) if remarks else ""
                })
                
                r += len(pallet_rows)
            else:
                r += 1
                
    print(f"  Parsed {len(pallets)} pallets from {sheet_name}.")
    return pallets

def main():
    wb = openpyxl.load_workbook("/Users/bigt_mbair/Library/CloudStorage/GoogleDrive-titankou2002@gmail.com/我的雲端硬碟/BT/Antigravity/倉庫全集/永安倉庫2025庫位表.xlsx", data_only=True)
    
    target_pattern = re.compile(r'^([A-C]-[A-H]區|花磚)$')
    sheets_to_parse = [name for name in wb.sheetnames if target_pattern.match(name)]
    
    print("Sheets to parse:", sheets_to_parse)
    
    all_pallets = []
    for name in sheets_to_parse:
        sheet = wb[name]
        sheet_pallets = parse_sheet(sheet, name)
        all_pallets.extend(sheet_pallets)
        
    print(f"\nTotal pallets parsed: {len(all_pallets)}")
    
    # Save as JSON
    out_path = "/Users/bigt_mbair/Library/CloudStorage/GoogleDrive-titankou2002@gmail.com/我的雲端硬碟/BT/Antigravity/倉庫全集/parsed_inventory.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(all_pallets, f, indent=2, ensure_ascii=False)
    print(f"Saved parsed inventory to: {out_path}")
    
    # Save as CSV
    import csv
    csv_path = "/Users/bigt_mbair/Library/CloudStorage/GoogleDrive-titankou2002@gmail.com/我的雲端硬碟/BT/Antigravity/倉庫全集/parsed_inventory.csv"
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "Sheet", "Slot", "Stack", "SKU", "Batch", "BoxQty", "PieceQty", 
            "BgColor", "FontColor", "Status", "IsLastPallet", "Remarks"
        ])
        writer.writeheader()
        writer.writerows(all_pallets)
    print(f"Saved CSV version to: {csv_path}")

if __name__ == "__main__":
    main()
